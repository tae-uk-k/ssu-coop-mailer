"""
core.py — 경로 · 설정 · 워크스페이스 · 파일 보관함 · 조사 · 명단

이 모듈은 UI를 전혀 모른다. 화면 쪽에서 함수만 불러 쓴다.
"""
from __future__ import annotations

import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# ──────────────────────────────────────────────────────────
# 경로
# ──────────────────────────────────────────────────────────

# 이 파일은 <프로젝트>/프로그램/ 안에 있다. 자료는 그 부모 폴더에 둔다.
# exe로 실행될 때는 exe 옆에 있어야 업데이트 후에도 설정이 남는다.
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).resolve().parent.parent

SETTINGS_DIR     = BASE_DIR / "설정"
WORKSPACES_DIR   = BASE_DIR / "행사별 작업"

ROOT_CFG_PATH    = SETTINGS_DIR / "config.json"
CREDENTIALS_PATH = SETTINGS_DIR / "credentials.json"
TOKEN_PATH       = SETTINGS_DIR / "token.json"


def ensure_dirs() -> None:
    """프로그램이 쓰는 폴더를 만들어 둔다."""
    SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
    WORKSPACES_DIR.mkdir(parents=True, exist_ok=True)

STATUS_COL = "발송상태"

# ──────────────────────────────────────────────────────────
# 설정 기본값
# ──────────────────────────────────────────────────────────

ROOT_CFG_DEFAULTS: dict = {
    "gmail_user":     "",
    "last_workspace": "",
    # 켤 때 새 코드를 확인할 주소 (manifest.json). 비우면 갱신하지 않는다.
    "update_url":     "",
}

WS_CFG_DEFAULTS: dict = {
    # 파일은 워크스페이스 안 상대경로로만 저장한다 (files/template_v1.pptx)
    "template_pptx":     "",
    "recipients_xlsx":   "",

    # 명단에서 앱이 반드시 알아야 하는 열
    "col_email":         "",

    # 제외 판정용 (선택)
    "col_category":      "",
    "skip_categories":   [],

    # PPT 자리 ↔ 명단 열 짝짓기  {"(업체명)": "업체명"}
    "slot_map":          {},
    "slot_josa":         True,
    # 본문에 원래 있는 괄호 — 짝짓지 않아도 경고하지 않는다
    "slot_ignore":       [],

    "file_name_pattern": "제안서 {{업체명}}",
    "email_subject":     "",
    "email_body":        "",
    "send_interval":     5,
}


def _read_json(path: Path, defaults: dict) -> dict:
    if path.exists():
        try:
            return {**defaults, **json.loads(path.read_text(encoding="utf-8"))}
        except Exception:
            pass
    return json.loads(json.dumps(defaults))  # 깊은 복사


def _write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def load_root_cfg() -> dict:
    """전역 설정을 읽는다.

    예전에는 SMTP로 보내느라 앱 비밀번호를 여기 평문으로 두었다. 지금은 OAuth만
    쓰므로 그 값과 옛 구조의 잔여 키를 지우고 파일도 다시 써 준다.
    """
    cfg = _read_json(ROOT_CFG_PATH, ROOT_CFG_DEFAULTS)
    extra = [k for k in cfg if k not in ROOT_CFG_DEFAULTS]
    if extra:
        for k in extra:
            cfg.pop(k, None)
        try:
            _write_json(ROOT_CFG_PATH, cfg)
        except Exception:
            pass
    return cfg


def save_root_cfg(data: dict) -> None:
    data = dict(data)
    data.pop("gmail_app_password", None)
    _write_json(ROOT_CFG_PATH, data)


def load_ws_cfg(name: str) -> dict:
    cfg = _read_json(ws_dir(name) / "config.json", WS_CFG_DEFAULTS)
    return _migrate_ws_cfg(name, cfg)


def save_ws_cfg(name: str, data: dict) -> None:
    _write_json(ws_dir(name) / "config.json", data)


def _migrate_ws_cfg(name: str, cfg: dict) -> dict:
    """구버전 설정을 새 구조로 옮긴다. 절대경로 → 보관함, 하드코딩 자리 → slot_map."""
    changed = False

    # 1) 절대경로로 저장돼 있던 파일을 보관함으로 들여온다
    for key, kind in (("template_pptx", "template"), ("recipients_xlsx", "recipients")):
        val = str(cfg.get(key, "") or "")
        if not val or val.startswith("files/"):
            continue
        src = Path(val)
        if src.is_absolute() and src.exists():
            try:
                cfg[key] = import_file(name, src, kind)
                changed = True
            except Exception:
                cfg[key] = ""
                changed = True
        else:
            cfg[key] = ""
            changed = True

    # 2) 예전 컬럼 설정 → 새 이름
    if not cfg.get("col_email") and cfg.get("col_email_legacy"):
        cfg["col_email"] = cfg.pop("col_email_legacy")
        changed = True

    # 3) 예전에는 (업체명)/(제품명)이 코드에 박혀 있었다. 열 설정이 남아 있으면 자리로 옮긴다
    if not cfg.get("slot_map"):
        legacy_company  = cfg.get("col_company", "")
        legacy_category = cfg.get("col_category", "")
        sm = {}
        if legacy_company:
            sm["(업체명)"] = legacy_company
            sm["(기업명)"] = legacy_company
        if legacy_category:
            sm["(제품명)"] = legacy_category
            sm["(협찬품)"] = legacy_category
        if sm:
            cfg["slot_map"] = sm
            changed = True

    if changed:
        try:
            save_ws_cfg(name, cfg)
        except Exception:
            pass
    return cfg


# ──────────────────────────────────────────────────────────
# 워크스페이스
# ──────────────────────────────────────────────────────────

def ws_dir(name: str) -> Path:
    return WORKSPACES_DIR / name


def files_dir(name: str) -> Path:
    return ws_dir(name) / "files"


def output_dir(name: str) -> Path:
    return ws_dir(name) / "제안서"


def list_workspaces() -> list[str]:
    if not WORKSPACES_DIR.exists():
        return []
    return sorted(d.name for d in WORKSPACES_DIR.iterdir()
                  if d.is_dir() and not d.name.startswith("."))


def create_workspace(name: str) -> bool:
    """폴더를 만든다. 이미 있으면 False."""
    d = ws_dir(name)
    if d.exists():
        return False
    files_dir(name).mkdir(parents=True)
    output_dir(name).mkdir(parents=True)
    save_ws_cfg(name, json.loads(json.dumps(WS_CFG_DEFAULTS)))
    return True


def delete_workspace(name: str) -> None:
    shutil.rmtree(ws_dir(name), ignore_errors=True)


def workspace_summary(name: str) -> dict:
    """첫 화면 '이어서 하기' 목록에 쓸 요약."""
    cfg = load_ws_cfg(name)
    out = output_dir(name)
    pdfs = len(list(out.glob("*.pdf"))) if out.exists() else 0
    done = 0
    try:
        if cfg.get("recipients_xlsx"):
            sheet = load_sheet(resolve(name, cfg["recipients_xlsx"]))
            done = sum(1 for r in sheet.rows if r.get(STATUS_COL, "") == "발송완료")
    except Exception:
        pass
    return {"name": name, "pdfs": pdfs, "sent": done, "step": ready_step(name, cfg)}


def ready_step(name: str, cfg: dict | None = None) -> int:
    """몇 단계까지 채워졌는지(0~5). 레일의 체크 표시에 쓴다."""
    cfg = cfg if cfg is not None else load_ws_cfg(name)
    done = 0
    if cfg.get("template_pptx") and cfg.get("recipients_xlsx"):
        done = 1
    else:
        return 0
    if cfg.get("col_email"):
        done = 2
    else:
        return 1
    slots = cfg.get("slot_map") or {}
    if slots and any(v for v in slots.values()):
        done = 3
    else:
        return 2
    if cfg.get("email_subject", "").strip() and cfg.get("email_body", "").strip():
        done = 4
    return done


# ──────────────────────────────────────────────────────────
# 파일 보관함 — 앱이 파일을 소유한다
# ──────────────────────────────────────────────────────────

_KIND_EXT = {"template": ".pptx", "recipients": ".xlsx"}


def import_file(ws: str, src: Path, kind: str) -> str:
    """올린 파일을 워크스페이스 안으로 복사하고 상대경로를 돌려준다.

    원본을 옮기거나 이름을 바꿔도 앱은 계속 쓸 수 있다.
    같은 종류를 다시 올리면 버전 번호가 올라가고 이전 버전은 남는다.
    """
    src = Path(src)
    if not src.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {src}")

    d = files_dir(ws)
    d.mkdir(parents=True, exist_ok=True)
    ext = _KIND_EXT.get(kind, src.suffix)
    ver = _next_version(ws, kind)
    dst = d / f"{kind}_v{ver}{ext}"
    shutil.copy2(src, dst)

    # 원본 이름을 따로 기억해 화면에 보여준다
    meta_path = d / "files.json"
    meta = {}
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            meta = {}
    meta[dst.name] = {
        "original": src.name,
        "imported": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    return f"files/{dst.name}"


def _next_version(ws: str, kind: str) -> int:
    d = files_dir(ws)
    if not d.exists():
        return 1
    best = 0
    for p in d.glob(f"{kind}_v*"):
        m = re.match(rf"{kind}_v(\d+)", p.stem)
        if m:
            best = max(best, int(m.group(1)))
    return best + 1


def file_versions(ws: str, kind: str) -> list[dict]:
    """최신순 버전 목록."""
    d = files_dir(ws)
    if not d.exists():
        return []
    meta = {}
    mp = d / "files.json"
    if mp.exists():
        try:
            meta = json.loads(mp.read_text(encoding="utf-8"))
        except Exception:
            pass
    out = []
    for p in d.glob(f"{kind}_v*"):
        m = re.match(rf"{kind}_v(\d+)", p.stem)
        if not m:
            continue
        info = meta.get(p.name, {})
        st = p.stat()
        out.append({
            "version":  int(m.group(1)),
            "rel":      f"files/{p.name}",
            "path":     p,
            "original": info.get("original", p.name),
            "imported": info.get("imported",
                                 datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")),
            "size":     st.st_size,
        })
    return sorted(out, key=lambda x: x["version"], reverse=True)


def resolve(ws: str, rel: str) -> Path:
    """상대경로를 실제 경로로. 예전 절대경로도 그대로 받아준다."""
    if not rel:
        return Path("")
    p = Path(rel)
    if p.is_absolute():
        return p
    return ws_dir(ws) / rel


def human_size(n: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024 or unit == "GB":
            return f"{n:.0f}{unit}" if unit == "B" else f"{n:.1f}{unit}"
        n /= 1024
    return f"{n:.1f}GB"


# ──────────────────────────────────────────────────────────
# 한글 조사 — PPT와 메일이 함께 쓴다
# ──────────────────────────────────────────────────────────

JOSA_MAP: dict[str, tuple[str, str]] = {
    "을를":   ("을", "를"),
    "이가":   ("이", "가"),
    "은는":   ("은", "는"),
    "와과":   ("과", "와"),
    "아야":   ("아", "야"),
    "으로로": ("으로", "로"),
    # 낱개로 써도 통하게
    "을":  ("을", "를"),
    "를":  ("을", "를"),
    "이":  ("이", "가"),
    "가":  ("이", "가"),
    "은":  ("은", "는"),
    "는":  ("은", "는"),
    "과":  ("과", "와"),
    "와":  ("과", "와"),
    "으로": ("으로", "로"),
    "로":  ("으로", "로"),
}

# 긴 것부터 맞춰야 '으로로'가 '으로'에 먹히지 않는다
_JOSA_KEYS = sorted(JOSA_MAP.keys(), key=len, reverse=True)


def has_batchim(ch: str) -> bool:
    if not ch:
        return False
    code = ord(ch)
    return 0xAC00 <= code <= 0xD7A3 and (code - 0xAC00) % 28 != 0


def pick_josa(word: str, key: str) -> str:
    """받침에 맞는 조사를 고른다. '으로'는 ㄹ받침이면 '로'."""
    with_b, no_b = JOSA_MAP[key]
    last = word[-1] if word else ""
    if not last:
        return no_b
    if key in ("으로로", "으로", "로"):
        code = ord(last)
        if 0xAC00 <= code <= 0xD7A3:
            jong = (code - 0xAC00) % 28
            return "로" if jong in (0, 8) else "으로"  # 받침 없음 또는 ㄹ
        return "로"
    return with_b if has_batchim(last) else no_b


def replace_placeholders(text: str, row: dict) -> str:
    """{{열}} 와 {{열은}} 을 값으로 바꾼다."""
    if not text:
        return text
    for col, val in row.items():
        if not col:
            continue
        val = "" if val is None else str(val)
        for key in _JOSA_KEYS:
            token = "{{" + col + key + "}}"
            if token in text:
                text = text.replace(token, val + pick_josa(val, key))
        text = text.replace("{{" + col + "}}", val)
    return text


def placeholder_names(text: str) -> set[str]:
    return set(re.findall(r"\{\{([^{}]+)\}\}", text or ""))


# ──────────────────────────────────────────────────────────
# 명단 (엑셀) — 앱 안에서 읽고 고친다
# ──────────────────────────────────────────────────────────

class Sheet:
    """엑셀 첫 시트를 표로 다룬다. 저장은 셀 단위로 제자리 수정해 서식을 지킨다."""

    def __init__(self, path: Path):
        self.path = Path(path)
        self.columns: list[str] = []
        self.rows: list[dict] = []      # 각 행에 '_row'(엑셀 행번호)가 들어 있다
        self._dirty: set[tuple[int, str]] = set()
        self._new_cols: list[str] = []
        self._load()

    def _load(self) -> None:
        wb = openpyxl.load_workbook(self.path, data_only=True)
        ws = wb.active
        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            headers.append("" if v is None else str(v).strip())
        # 뒤쪽 빈 헤더는 버린다
        while headers and not headers[-1]:
            headers.pop()
        self.columns = headers

        for r in range(2, ws.max_row + 1):
            vals = {}
            empty = True
            for i, h in enumerate(headers, start=1):
                v = ws.cell(r, i).value
                s = "" if v is None else str(v).strip()
                if s:
                    empty = False
                vals[h] = s
            if empty:
                continue
            vals["_row"] = r
            self.rows.append(vals)
        wb.close()

    # ── 편집 ──
    def set(self, idx: int, col: str, value: str) -> None:
        if idx < 0 or idx >= len(self.rows):
            return
        if self.rows[idx].get(col, "") == value:
            return
        self.rows[idx][col] = value
        self._dirty.add((idx, col))

    def ensure_column(self, name: str) -> None:
        if name and name not in self.columns:
            self.columns.append(name)
            self._new_cols.append(name)
            for r in self.rows:
                r.setdefault(name, "")

    def add_row(self) -> int:
        last = max((r["_row"] for r in self.rows), default=1)
        new = {c: "" for c in self.columns}
        new["_row"] = last + 1
        self.rows.append(new)
        for c in self.columns:
            self._dirty.add((len(self.rows) - 1, c))
        return len(self.rows) - 1

    @property
    def dirty(self) -> bool:
        return bool(self._dirty) or bool(self._new_cols)

    def save(self) -> None:
        if not self.dirty:
            return
        wb = openpyxl.load_workbook(self.path)
        ws = wb.active
        col_at = _header_index(ws)
        for name in self._new_cols:
            if name not in col_at:
                nc = ws.max_column + 1
                ws.cell(1, nc).value = name
                col_at[name] = nc
        for idx, col in self._dirty:
            if idx >= len(self.rows):
                continue
            ci = col_at.get(col)
            if ci is None:
                ci = ws.max_column + 1
                ws.cell(1, ci).value = col
                col_at[col] = ci
            ws.cell(self.rows[idx]["_row"], ci).value = self.rows[idx].get(col, "")
        wb.save(self.path)
        wb.close()
        self._dirty.clear()
        self._new_cols.clear()


def _header_index(ws) -> dict[str, int]:
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None and str(v).strip():
            out[str(v).strip()] = c
    return out


def load_sheet(path: Path) -> Sheet:
    return Sheet(path)


def guess_column(columns: list[str], *hints: str) -> str:
    """열 이름을 알아서 골라 준다. 정확히 같은 것 → 부분 일치 순."""
    for h in hints:
        for c in columns:
            if c == h:
                return c
    for h in hints:
        for c in columns:
            if h and (h in c or c in h):
                return c
    return ""


EMAIL_HINTS    = ("이메일", "메일", "email", "E-mail", "주소")
COMPANY_HINTS  = ("업체명", "기업명", "회사명", "회사", "기업", "업체", "브랜드")
CATEGORY_HINTS = ("작은 유형", "주력 카테고리", "카테고리", "품목", "제품", "유형", "분류")


# ── 발송 대상 판정 ────────────────────────────────────────

SEND        = "send"
NO_EMAIL    = "no_email"
BAD_EMAIL   = "bad_email"
ALREADY     = "already"
SKIP_CAT    = "skip_category"
BOUNCED     = "bounced"

REASON_TEXT = {
    SEND:      "보낼 예정",
    NO_EMAIL:  "이메일 없음",
    BAD_EMAIL: "이메일 형식 오류",
    ALREADY:   "이미 보냄",
    SKIP_CAT:  "제외 품목",
    BOUNCED:   "반송된 주소",
}

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def classify(sheet: Sheet, cfg: dict, log: "SendLog | None" = None) -> list[dict]:
    """행마다 보낼지 말지와 그 이유를 붙인다. 보내기 전에 미리 보여주기 위한 것.

    log 를 주면 엑셀의 '발송상태' 열보다 그쪽을 먼저 본다. 명단을 바꿔도
    어디까지 보냈는지가 남아 있어야 하기 때문이다.
    """
    col_em  = cfg.get("col_email", "")
    col_cat = cfg.get("col_category", "")
    skips   = {str(s).strip() for s in (cfg.get("skip_categories") or []) if str(s).strip()}
    name_col = _display_name_column(sheet, cfg)

    out = []
    for i, row in enumerate(sheet.rows):
        email  = (row.get(col_em, "") or "").strip()
        status = (log.status_of(email) if log else "")             or (row.get(STATUS_COL, "") or "").strip()
        cat    = (row.get(col_cat, "") or "").strip() if col_cat else ""

        if status == "발송완료":
            kind = ALREADY
        elif status == "반송됨":
            kind = BOUNCED
        elif not email or email in ("미확인", "-"):
            kind = NO_EMAIL
        elif not _EMAIL_RE.match(re.sub(r"\s+", "", email.split("/")[0])):
            kind = BAD_EMAIL
        elif cat and cat in skips:
            kind = SKIP_CAT
        else:
            kind = SEND

        out.append({
            "index":   i,
            "row":     row,
            "when":    log.when_of(email) if log else "",
            "name":    (row.get(name_col, "") or "").strip() if name_col else f"{i + 1}행",
            "email":   email,
            "category": cat,
            "kind":    kind,
            "reason":  REASON_TEXT[kind],
        })
    return out


def _display_name_column(sheet: Sheet, cfg: dict) -> str:
    """표에 기업 이름으로 보여줄 열. 자리 짝짓기에 쓰인 열을 우선 쓴다."""
    for slot, col in (cfg.get("slot_map") or {}).items():
        if col and col in sheet.columns and any(h in slot for h in ("업체", "기업", "회사")):
            return col
    return guess_column(sheet.columns, *COMPANY_HINTS) or (sheet.columns[0] if sheet.columns else "")


def first_email(raw: str) -> str:
    """'a@b.com / c@d.com' 처럼 여러 개면 첫 번째만 쓴다."""
    return re.sub(r"\s+", "", str(raw or "").split("/")[0]).strip()


def safe_filename(name: str) -> str:
    name = re.sub(r'[\\/:*?"<>|]', "_", str(name or "")).strip()
    name = re.sub(r"\s+", " ", name).strip(" .")
    return name[:150] or "제안서"


def unique_name(base: str, used: set[str]) -> str:
    """이미 쓴 이름이면 뒤에 번호를 붙인다.

    이름 규칙에 기업마다 달라지는 값이 없으면 파일이 서로 덮어써진다.
    화면에서 미리 경고하지만, 만에 하나를 대비해 여기서도 막는다.
    """
    name, n = base, 2
    while name.lower() in used:
        name = f"{base}_{n}"
        n += 1
    used.add(name.lower())
    return name


def preview_filenames(pattern: str, rows: list[dict], limit: int = 500,
                      email_col: str = "", name_col: str = "") -> dict:
    """이름 규칙이 기업마다 다른 이름을 만드는지 미리 따져 본다.

    이름이 겹칠 때 원인은 둘 중 하나이고, 고칠 방법이 서로 다르다.
      · 규칙에 기업마다 달라지는 값이 아예 없다   → 규칙을 고쳐야 한다
      · 규칙은 멀쩡한데 명단에 같은 값이 여러 번   → 명단을 고쳐야 한다
    """
    sample = rows[:limit]
    names = [safe_filename(replace_placeholders(pattern, r)) for r in sample]

    left = set()
    for n in names:
        left |= placeholder_names(n)

    # 겹치는 이름 — 어느 기업이 몇 곳인지, 이메일도 같은지까지 본다.
    #   같은 이메일  → 명단에 잘못 들어간 중복. 그대로 두면 같은 곳에 여러 번 간다
    #   다른 이메일  → 담당자가 여러 분. 일부러 그런 것일 수 있다
    where: dict[str, list[int]] = {}
    for i, n in enumerate(names):
        where.setdefault(n, []).append(i)

    dupes = []
    for n, idxs in where.items():
        if len(idxs) < 2:
            continue
        mails = []
        if email_col:
            for i in idxs:
                m = first_email(sample[i].get(email_col, "")).lower()
                if m and m not in mails:
                    mails.append(m)
        dupes.append({
            "name":       n,
            "count":      len(idxs),
            "company":    (sample[idxs[0]].get(name_col, "") or "").strip()
                          if name_col else "",
            "emails":     mails,
            "same_email": len(mails) == 1,
        })
    dupes.sort(key=lambda d: -d["count"])

    # 규칙에 열이 하나라도 들어 있나 (조사가 붙은 것도 인정)
    has_var = bool(placeholder_names(pattern))

    return {
        "names":    names,
        "first":    names[0] if names else "",
        "unique":   len(set(names)),
        "total":    len(names),
        "leftover": sorted(left),          # 엑셀에 없는 열을 적었을 때
        "dupes":    dupes,                 # [(이름, 몇 곳), …]
        "has_var":  has_var,               # 규칙에 {{열}} 이 들어 있나
    }


# ──────────────────────────────────────────────────────────
# 보내기 전 점검 — 안 바뀐 채 나갈 것을 찾는다
# ──────────────────────────────────────────────────────────

# 자리 이름에 이런 말이 들어 있으면 '바뀌어야 할 곳' 으로 본다.
# (스태프) (강원도 원주) 같은 본문 괄호와 구분하기 위한 것.
PLACEHOLDER_HINTS = (
    "기업", "회사", "업체", "브랜드", "사명", "상호",
    "제품", "품목", "상품", "협찬", "물품", "지원",
    "담당자", "성함", "이름", "직함", "부서",
)


def token_column(token: str, columns: list[str]) -> str:
    """{{업체명은}} 처럼 조사가 붙어도 어느 열인지 찾아 준다. 없으면 빈 문자열."""
    if token in columns:
        return token
    for key in _JOSA_KEYS:
        if token.endswith(key):
            base = token[: -len(key)]
            if base in columns:
                return base
    return ""


def unmatched_tokens(text: str, columns: list[str]) -> list[str]:
    """{{...}} 중 명단에 없는 것. 이대로 보내면 글자가 그대로 나간다."""
    out = []
    for t in sorted(placeholder_names(text)):
        if not token_column(t, columns):
            out.append(t)
    return out


def looks_like_placeholder(slot_inner: str, columns: list[str]) -> bool:
    """PPT 의 (자리) 가 '원래 바뀌어야 할 곳' 처럼 보이는지."""
    if guess_column(columns, slot_inner):
        return True
    return any(h in slot_inner for h in PLACEHOLDER_HINTS)


def pre_send_issues(cfg: dict, slots, sheet) -> tuple[list[dict], list[dict]]:
    """(꼭 고쳐야 할 것, 확인만 하면 되는 것) 을 돌려준다.

    돌려주는 각 항목: {"where", "token", "detail", "how"}
    """
    must, check = [], []
    cols = list(sheet.columns) if sheet else []

    # 1. 메일 제목·본문·파일이름 — 명단에 없는 {{토큰}}
    for where, text in (("메일 제목", cfg.get("email_subject", "")),
                        ("메일 본문", cfg.get("email_body", "")),
                        ("파일 이름", cfg.get("file_name_pattern", ""))):
        for t in unmatched_tokens(text, cols):
            must.append({
                "where": where,
                "token": "{{" + t + "}}",
                "detail": f"명단에 '{t}' 열이 없어요.",
                "how": ("이 글자가 그대로 나갑니다. "
                        + (f"'{guess_column(cols, t)}' 열을 쓰시려던 것 같아요."
                           if guess_column(cols, t) else "명단의 열 이름을 확인해 주세요.")),
            })

    # 2. 제안서 PPT — 짝짓지 않은 자리
    ignore = set(cfg.get("slot_ignore") or [])
    smap = {k: v for k, v in (cfg.get("slot_map") or {}).items() if v}
    for s in (slots or []):
        if s.name in smap or s.name in ignore:
            continue
        item = {
            "where": "제안서",
            "token": s.name,
            "detail": f"{s.where} 에 있어요.",
            "how": "3단계에서 넣을 열을 고르거나, 원래 그런 문구면 '그냥 두기' 를 눌러 주세요.",
        }
        (must if looks_like_placeholder(s.inner, cols) else check).append(item)

    return must, check


# ──────────────────────────────────────────────────────────
# 발송 기록 — 엑셀이 아니라 행사 폴더에 남긴다
# ──────────────────────────────────────────────────────────

class SendLog:
    """어디까지 보냈는지를 이메일 주소로 기억한다.

    예전에는 엑셀의 '발송상태' 열에만 적었다. 그런데 앱은 올린 엑셀의
    '사본'을 쓰기 때문에, 원본을 고쳐 다시 올리면 그 열이 없는 새 파일이
    되어 기록이 통째로 사라졌다. 명단을 바꿔도 남아야 하는 정보라
    행사 폴더에 따로 둔다.

    엑셀에도 계속 적어 준다 — 사람이 열어 볼 수 있어야 하니까.
    """

    FILE = "발송기록.json"

    def __init__(self, ws: str):
        self.ws = ws
        self.path = ws_dir(ws) / self.FILE
        self.rows: dict[str, dict] = {}
        self._load()

    def _load(self) -> None:
        try:
            data = json.loads(self.path.read_text(encoding="utf-8"))
            self.rows = data.get("보낸곳") or {}
        except Exception:
            self.rows = {}

    def save(self) -> None:
        try:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            self.path.write_text(
                json.dumps({"보낸곳": self.rows}, ensure_ascii=False, indent=2),
                encoding="utf-8")
        except Exception:
            pass

    @staticmethod
    def _key(email: str) -> str:
        return first_email(email).lower()

    def mark(self, email: str, name: str, status: str) -> None:
        k = self._key(email)
        if not k:
            return
        self.rows[k] = {
            "이름": name,
            "상태": status,
            "때":  datetime.now().strftime("%Y-%m-%d %H:%M"),
        }

    def status_of(self, email: str) -> str:
        return (self.rows.get(self._key(email)) or {}).get("상태", "")

    def when_of(self, email: str) -> str:
        return (self.rows.get(self._key(email)) or {}).get("때", "")

    def clear(self, email: str) -> None:
        self.rows.pop(self._key(email), None)

    def count(self, status: str) -> int:
        return sum(1 for v in self.rows.values() if v.get("상태") == status)

    # ── 명단과 맞추기 ──
    def absorb(self, sheet, email_col: str) -> int:
        """엑셀에 적혀 있던 상태를 기록으로 들여온다.

        예전 버전 명단이나, 손으로 표시해 둔 것을 살리기 위한 것.
        """
        if not email_col:
            return 0
        n = 0
        for row in sheet.rows:
            st = (row.get(STATUS_COL, "") or "").strip()
            if st and not self.status_of(row.get(email_col, "")):
                k = self._key(row.get(email_col, ""))
                if k:
                    self.rows[k] = {"이름": "", "상태": st, "때": ""}
                    n += 1
        if n:
            self.save()
        return n

    def apply_to(self, sheet, email_col: str) -> int:
        """기록을 엑셀 '발송상태' 열에 다시 써 넣는다.

        새 명단을 올려도 사람이 엑셀만 열어 봐도 알 수 있게 한다.
        """
        if not email_col or not self.rows:
            return 0
        sheet.ensure_column(STATUS_COL)
        n = 0
        for i, row in enumerate(sheet.rows):
            st = self.status_of(row.get(email_col, ""))
            if st and (row.get(STATUS_COL, "") or "").strip() != st:
                sheet.set(i, STATUS_COL, st)
                n += 1
        return n
